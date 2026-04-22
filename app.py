"""PDF Crawler — scan folder · match Excel list · copy most-recent PDFs"""

import os
import sys
import json
import shutil
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from datetime import datetime
import functools
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
from openpyxl import load_workbook

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False


# ── App Directory (handle frozen executable) ──────────────────────────────────
if getattr(sys, 'frozen', False):
    # Running as compiled executable
    APP_DIR = Path(sys.executable).parent
else:
    # Running as script
    APP_DIR = Path(__file__).parent


# ── Palette ───────────────────────────────────────────────────────────────────
C = {
    "bg":      "#0f172a",
    "surface": "#1e293b",
    "border":  "#334155",
    "primary": "#6366f1",
    "hover":   "#4f46e5",
    "accent":  "#22d3ee",
    "success": "#4ade80",
    "error":   "#f87171",
    "warn":    "#fbbf24",
    "text":    "#f1f5f9",
    "muted":   "#94a3b8",
    "input":   "#0f172a",
}

CACHE_DIR = APP_DIR / "cache"


def make_cache_path():
    """Return a timestamped path inside the project cache folder."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return str(CACHE_DIR / f"scan_{ts}.json")


# ══════════════════════════════════════════════════════════════════════════════
# CORE – pure functions
# ══════════════════════════════════════════════════════════════════════════════

def scan_pdfs(folder, max_workers=4):
    """Return ({lowercase_filename: [full_paths]}, {path: creation_time}) using parallel scanning."""
    result = {}
    file_times = {}
    
    def process_pdfs(pdf_paths):
        """Process a batch of PDF paths."""
        batch_result = {}
        batch_times = {}
        for pdf_path in pdf_paths:
            name = pdf_path.name.lower()
            batch_result.setdefault(name, []).append(str(pdf_path))
            try:
                batch_times[str(pdf_path)] = os.path.getctime(str(pdf_path))
            except OSError:
                batch_times[str(pdf_path)] = 0
        return batch_result, batch_times
    
    # Collect all PDF paths using pathlib (filters only .pdf files)
    all_pdfs = list(Path(folder).rglob('*.pdf'))
    
    # Process in parallel batches
    if all_pdfs:
        batch_size = max(1, len(all_pdfs) // max_workers)
        batches = [all_pdfs[i:i + batch_size] for i in range(0, len(all_pdfs), batch_size)]
        
        with ThreadPoolExecutor(max_workers=max_workers) as exe:
            for batch_result, batch_times in exe.map(process_pdfs, batches):
                for name, paths in batch_result.items():
                    result.setdefault(name, []).extend(paths)
                file_times.update(batch_times)
    
    return result, file_times


def save_cache(cache, path):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(cache, fh, indent=2, ensure_ascii=False)


def load_cache(path):
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


CONFIG_FILE = APP_DIR / "config.json"


def load_config():
    """Load app config. Currently unused but kept for compatibility."""
    return {}


def save_config():
    """Save app config. Currently unused but kept for compatibility."""
    pass


def read_filenames(file_path):
    """Read filenames from Sheet2 (1st sheet), column B, starting from row 4.
    
    Uses openpyxl with data_only=True for faster parsing.
    Only reads the needed column.
    """
    if file_path.lower().endswith(".csv"):
        raise ValueError("Only Excel files are supported. Expected .xlsx or .xls")
    
    try:
        # Use openpyxl for faster parsing (data_only skips formulas)
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        
        # Read only column B (column 2), starting from row 4
        filenames = []
        for row_idx in range(4, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=2).value
            if cell_value is not None:
                filenames.append(str(cell_value).strip())
        
        wb.close()
        
        if not filenames:
            raise ValueError("No filenames found in Sheet2, Column B, starting from Row 4")
        
        return filenames
    except Exception as e:
        # Fallback to pandas if openpyxl fails
        df = pd.read_excel(file_path, sheet_name=0, header=None, 
                          usecols=[1], skiprows=3)
        filenames = (
            df.iloc[:, 0]
            .dropna()
            .astype(str)
            .str.strip()
            .tolist()
        )
        
        if not filenames:
            raise ValueError("No filenames found in Sheet2, Column B, starting from Row 4")
        
        return filenames


def match_filenames(names, cache):
    """Return (found_dict, missing_list).
    
    Builds a fast search index for substring matching.
    Example: searching for '2222' will match '2222.pdf', '2222A.pdf', '2222_0.pdf', etc.
    """
    # Build search index once (fast O(n) preprocessing)
    search_index = {}
    for pdf_filename, paths in cache.items():
        pdf_name_no_ext = pdf_filename[:-4] if pdf_filename.endswith(".pdf") else pdf_filename
        paths_list = list(paths) if isinstance(paths, (list, tuple)) else [paths]
        
        # Add exact match
        search_index[pdf_name_no_ext] = paths_list
        
        # Add all substring matches (1-char and longer)
        for i in range(len(pdf_name_no_ext)):
            for j in range(i + 1, len(pdf_name_no_ext) + 1):
                substring = pdf_name_no_ext[i:j]
                if substring not in search_index:
                    search_index[substring] = []
                search_index[substring].extend(paths_list)
    
    found, missing = {}, []
    for name in names:
        search_term = name.lower().rstrip('.pdf')
        
        # Fast O(1) lookup in index
        if search_term in search_index:
            found[name] = search_index[search_term]
        else:
            missing.append(name)
    
    return found, missing


def most_recent_path(paths, file_times=None):
    """Return the path with the most recent creation time.
    
    Uses cached file_times if provided (much faster).
    Falls back to os.path.getctime if cache not available.
    """
    if not paths:
        return None
    
    if file_times:
        # Use cached times (no I/O)
        return max(paths, key=lambda p: file_times.get(p, 0))
    else:
        # Fallback: fetch from file system
        return max(paths, key=os.path.getctime)


def copy_matches(matches, dest, file_times=None, max_workers=4):
    """Copy the most-recent PDF for each match into dest. Return (copied, errors).
    
    Uses parallel threads for faster copying.
    """
    Path(dest).mkdir(parents=True, exist_ok=True)
    
    def copy_one(item):
        """Copy a single match."""
        name, paths = item
        try:
            src = most_recent_path(paths, file_times)
            if not src:
                return (name, None, None, "No paths available")
            dst = os.path.join(dest, os.path.basename(src))
            shutil.copy2(src, dst)
            return (name, src, dst, None)
        except Exception as exc:
            return (name, None, None, str(exc))
    
    # Process copies in parallel
    copied, errors = [], []
    with ThreadPoolExecutor(max_workers=max_workers) as exe:
        for name, src, dst, error in exe.map(copy_one, matches.items()):
            if error is None:
                copied.append((name, src, dst))
            else:
                errors.append((name, error))
    
    return copied, errors


# ══════════════════════════════════════════════════════════════════════════════
# WIDGET HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _divider(parent):
    tk.Frame(parent, bg=C["border"], height=1).pack(fill="x", padx=28, pady=(10, 14))


def _section_label(parent, num, title, sub=""):
    f = tk.Frame(parent, bg=C["bg"])
    f.pack(fill="x", pady=(10, 2))
    tk.Label(f, text=f" {num} ", bg=C["primary"], fg=C["text"],
             font=("Segoe UI", 9, "bold"), padx=4, pady=1).pack(side="left")
    tk.Label(f, text=f"  {title}", bg=C["bg"], fg=C["text"],
             font=("Segoe UI", 11, "bold")).pack(side="left")
    if sub:
        tk.Label(f, text=f"  {sub}", bg=C["bg"], fg=C["muted"],
                 font=("Segoe UI", 9)).pack(side="left", pady=2)


def _label(parent, text):
    return tk.Label(parent, text=text, bg=C["bg"], fg=C["muted"],
                    font=("Segoe UI", 10))


def _entry(parent, var, width=None):
    kw = dict(
        textvariable=var,
        bg=C["input"], fg=C["text"],
        insertbackground=C["text"],
        font=("Segoe UI", 10),
        relief="flat", bd=8,
        highlightthickness=1,
        highlightbackground=C["border"],
        highlightcolor=C["primary"],
    )
    if width:
        kw["width"] = width
    return tk.Entry(parent, **kw)


def _button(parent, text, cmd, primary=False, large=False, small=False):
    bg  = C["primary"] if primary else C["surface"]
    hov = C["hover"]   if primary else C["border"]
    py  = 10 if large else (3 if small else 6)
    px  = 22 if large else (8 if small else 14)
    fnt = ("Segoe UI", 11, "bold") if large else ("Segoe UI", 10)
    b = tk.Button(parent, text=text, command=cmd,
                  bg=bg, fg=C["text"], activebackground=hov,
                  activeforeground=C["text"], relief="flat",
                  padx=px, pady=py, font=fnt, cursor="hand2", bd=0)
    b.bind("<Enter>", lambda _e: b.configure(bg=hov))
    b.bind("<Leave>", lambda _e: b.configure(bg=bg))
    return b


# ══════════════════════════════════════════════════════════════════════════════
# APP CLASS – UI wiring only, all logic lives in pure functions above
# ══════════════════════════════════════════════════════════════════════════════

class PDFCrawler:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF Crawler")
        self.root.configure(bg=C["bg"])
        self.root.geometry("820x800")
        self.root.minsize(680, 680)

        self.var_source = tk.StringVar()
        self.var_dest   = tk.StringVar()
        self.var_excel  = tk.StringVar()

        self._cache_map  = None   # currently active pdf map
        self._cache_path = None   # path to the active json file
        self._file_times = None   # cached file creation times for fast lookup

        self._build()
        
        # Auto-load most recent cache on startup (much faster than rescanning)
        self._auto_load_recent_cache()

    # ── layout ────────────────────────────────────────────────────────────────

    def _build(self):
        # Header
        hdr = tk.Frame(self.root, bg=C["bg"])
        hdr.pack(fill="x", padx=28, pady=(22, 0))
        tk.Label(hdr, text="PDF Crawler", font=("Segoe UI", 20, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(side="left")
        tk.Label(hdr, text="— scan · match · copy",
                 font=("Segoe UI", 11), bg=C["bg"], fg=C["muted"]).pack(
                 side="left", padx=10, pady=4)

        _divider(self.root)

        body = tk.Frame(self.root, bg=C["bg"])
        body.pack(fill="both", expand=False, padx=28)

        # Step 1 — Source folder
        _section_label(body, "1", "Source Folder",
                       "Scanned recursively for PDF files")
        r1 = tk.Frame(body, bg=C["bg"])
        r1.pack(fill="x", pady=(4, 6))
        _entry(r1, self.var_source).pack(side="left", fill="x", expand=True)
        _button(r1, "Browse", self._browse_source).pack(side="left", padx=(6, 0))
        _button(r1, "Scan PDFs →", self._do_scan, primary=True).pack(
            side="left", padx=(6, 0))

        rc = tk.Frame(body, bg=C["surface"],
                      highlightthickness=1, highlightbackground=C["border"])
        rc.pack(fill="x", pady=(0, 14))
        tk.Label(rc, text="Cache:", bg=C["surface"], fg=C["muted"],
                 font=("Segoe UI", 9), padx=8, pady=6).pack(side="left")
        self.cache_lbl = tk.Label(rc, text="  No cache loaded",
                                  bg=C["surface"], fg=C["muted"],
                                  font=("Consolas", 9), anchor="w", pady=6)
        self.cache_lbl.pack(side="left", fill="x", expand=True)
        _button(rc, "Load Cache", self._load_cache_file,
                small=True).pack(side="right", padx=8, pady=4)

        # Step 2 — Excel / CSV
        _section_label(body, "2", "Excel / CSV File",
                       "Drag & drop or use Browse")
        self._build_drop_zone(body)

        # Step 3 — Destination
        _section_label(body, "3", "Destination Folder",
                       "Where matched PDFs will be copied")
        r3 = tk.Frame(body, bg=C["bg"])
        r3.pack(fill="x", pady=(4, 14))
        _entry(r3, self.var_dest).pack(side="left", fill="x", expand=True)
        _button(r3, "Browse", self._browse_dest).pack(side="left", padx=(6, 0))

        # Run button
        run_wrap = tk.Frame(body, bg=C["bg"])
        run_wrap.pack(fill="x", pady=(4, 2))
        _button(run_wrap, "⚡  Run — Match & Copy PDFs",
                self._do_run, primary=True, large=True).pack(fill="x")

        _divider(self.root)

        # Log area
        log_hdr = tk.Frame(self.root, bg=C["bg"])
        log_hdr.pack(fill="x", padx=28)
        _label(log_hdr, "Log").pack(side="left")
        _button(log_hdr, "Clear", self._clear_log, small=True).pack(side="right")

        log_wrap = tk.Frame(self.root, bg=C["surface"],
                            highlightthickness=1,
                            highlightbackground=C["border"])
        log_wrap.pack(fill="both", expand=True, padx=28, pady=(4, 20))

        self.log_box = tk.Text(
            log_wrap, bg=C["surface"], fg=C["text"],
            font=("Consolas", 9), relief="flat", bd=10,
            wrap="word", state="disabled",
            selectbackground=C["primary"])

        sb = ttk.Scrollbar(log_wrap, orient="vertical",
                           command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.log_box.pack(fill="both", expand=True)

        for tag, color in (
            ("ok",   C["success"]),
            ("err",  C["error"]),
            ("warn", C["warn"]),
            ("dim",  C["muted"]),
            ("hi",   C["accent"]),
        ):
            self.log_box.tag_config(tag, foreground=color)

    def _build_drop_zone(self, parent):
        frame = tk.Frame(parent, bg=C["surface"],
                         highlightthickness=1,
                         highlightbackground=C["border"])
        frame.pack(fill="x", pady=(4, 0))
        inner = tk.Frame(frame, bg=C["surface"])
        inner.pack(fill="x", padx=2, pady=2)

        self.drop_lbl = tk.Label(
            inner, text="  Drop Excel / CSV file here …",
            bg=C["surface"], fg=C["muted"],
            font=("Segoe UI", 10), anchor="w", pady=10)
        self.drop_lbl.pack(side="left", fill="x", expand=True)

        _button(inner, "Browse", self._browse_excel,
                small=True).pack(side="right", padx=8, pady=6)

        if HAS_DND:
            for widget in (frame, inner, self.drop_lbl):
                widget.drop_target_register(DND_FILES)
                widget.dnd_bind("<<Drop>>", self._on_drop)

    # ── actions ───────────────────────────────────────────────────────────────

    def _browse_source(self):
        p = filedialog.askdirectory()
        if p:
            self.var_source.set(p)

    def _browse_dest(self):
        p = filedialog.askdirectory()
        if p:
            self.var_dest.set(p)

    def _browse_excel(self):
        p = filedialog.askopenfilename(
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("All", "*.*")])
        if p:
            self._set_excel(p)

    def _on_drop(self, event):
        path = event.data.strip().strip("{}")
        if os.path.isfile(path) and path.lower().endswith((".xlsx", ".xls", ".csv")):
            self._set_excel(path)
        else:
            self._log("Only .xlsx / .xls / .csv files accepted.", "warn")

    def _set_excel(self, path):
        self.var_excel.set(path)
        self.drop_lbl.configure(
            text=f"  ✓  {os.path.basename(path)}", fg=C["accent"])

    def _do_scan(self):
        folder = self.var_source.get().strip()
        if not os.path.isdir(folder):
            messagebox.showerror("Invalid path", "Select a valid source folder first.")
            return
        self._log(f"Scanning {folder} …", "dim")

        def worker():
            try:
                cache_path = make_cache_path()
                pdf_map, file_times = scan_pdfs(folder)  # Now returns both
                save_cache(pdf_map, cache_path)
                self._cache_map  = pdf_map
                self._cache_path = cache_path
                self._file_times = file_times
                unique = len(pdf_map)
                total  = sum(len(v) for v in pdf_map.values())
                self._post(
                    f"Scan complete — {unique} unique names, "
                    f"{total} total files.  Cache saved → {cache_path}", "ok")
                self.root.after(0, self._update_cache_label)
            except Exception as exc:
                self._post(f"Scan error: {exc}", "err")

        threading.Thread(target=worker, daemon=True).start()

    def _do_run(self):
        src  = self.var_source.get().strip()
        dst  = self.var_dest.get().strip()
        xlsx = self.var_excel.get().strip()

        if not os.path.isdir(src):
            messagebox.showerror("Invalid path", "Select a valid source folder.")
            return
        if not os.path.isfile(xlsx):
            messagebox.showerror("Missing file",  "Select an Excel / CSV file.")
            return
        if not dst:
            messagebox.showerror("Missing path",  "Select a destination folder.")
            return

        def worker():
            try:
                # Use active cache or auto-scan
                if self._cache_map is not None:
                    pdf_map = self._cache_map
                    file_times = self._file_times
                    self._post(
                        f"Using cache: {os.path.basename(self._cache_path)}  "
                        f"({len(pdf_map)} unique names)", "dim")
                else:
                    self._post("No cache loaded — scanning first …", "warn")
                    cache_path = make_cache_path()
                    pdf_map, file_times = scan_pdfs(src)
                    save_cache(pdf_map, cache_path)
                    self._cache_map  = pdf_map
                    self._cache_path = cache_path
                    self._file_times = file_times
                    unique = len(pdf_map)
                    total  = sum(len(v) for v in pdf_map.values())
                    self._post(
                        f"Scanned: {unique} unique / {total} total PDFs.  "
                        f"Cache saved → {cache_path}", "ok")
                    self.root.after(0, self._update_cache_label)

                # Read spreadsheet
                self._post(
                    f"Reading filenames from Sheet2, Column B (starting Row 4) in "
                    f"{os.path.basename(xlsx)} …", "dim")
                names = read_filenames(xlsx)
                self._post(f"{len(names)} file names found in spreadsheet.", "hi")

                # Match
                found, missing = match_filenames(names, pdf_map)
                self._post(
                    f"Matched: {len(found)}  |  Not found: {len(missing)}")
                for m in missing:
                    self._post(f"    NOT FOUND: {m}", "warn")

                # Copy with parallel workers and file_times cache
                if found:
                    self._post(f"Copying {len(found)} PDF(s) → {dst} …", "dim")
                    copied, errors = copy_matches(found, dst, file_times, max_workers=4)
                    for _name, src_path, _dst_path in copied:
                        self._post(f"    ✓  {os.path.basename(src_path)}", "ok")
                    for name, err in errors:
                        self._post(f"    ✗  {name}: {err}", "err")
                    self._post(
                        f"Done — {len(copied)} copied, "
                        f"{len(errors)} errors, {len(missing)} not found.", "ok")
                else:
                    self._post("No matches found — nothing copied.", "warn")

            except Exception as exc:
                self._post(f"Error: {exc}", "err")

        threading.Thread(target=worker, daemon=True).start()

    def _load_cache_file(self):
        initial = str(CACHE_DIR) if CACHE_DIR.exists() else str(APP_DIR)
        path = filedialog.askopenfilename(
            title="Select cache JSON",
            initialdir=initial,
            filetypes=[("JSON cache", "*.json"), ("All files", "*.*")])
        if not path:
            return
        try:
            data = load_cache(path)
            if data is None:
                raise ValueError("File could not be read.")
            self._cache_map  = data
            self._cache_path = path
            self._update_cache_label()
            self._log(
                f"Cache loaded: {os.path.basename(path)}  "
                f"({len(data)} unique names)", "ok")
        except Exception as exc:
            messagebox.showerror("Load error", str(exc))

    def _update_cache_label(self):
        if self._cache_path:
            name  = os.path.basename(self._cache_path)
            count = len(self._cache_map) if self._cache_map else 0
            self.cache_lbl.configure(
                text=f"  {name}  ({count} entries)", fg=C["accent"])
        else:
            self.cache_lbl.configure(text="  No cache loaded", fg=C["muted"])

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _log(self, msg, tag=""):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"[{ts}]  {msg}\n", tag)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _post(self, msg, tag=""):
        """Thread-safe log."""
        self.root.after(0, lambda: self._log(msg, tag))

    def _auto_load_recent_cache(self):
        """Auto-load the most recent cache file on startup (much faster than rescanning)."""
        if not CACHE_DIR.exists():
            return
        
        try:
            # Find most recent cache file
            cache_files = sorted(
                CACHE_DIR.glob('scan_*.json'),
                key=lambda p: p.stat().st_mtime,
                reverse=True
            )
            
            if not cache_files:
                return
            
            recent_cache = cache_files[0]
            data = load_cache(str(recent_cache))
            
            if data:
                self._cache_map  = data
                self._cache_path = str(recent_cache)
                # Rebuild file_times from cache (fast operation)
                self._file_times = {}
                for paths in data.values():
                    for path in paths:
                        try:
                            self._file_times[path] = os.path.getctime(path)
                        except OSError:
                            pass
                
                self.root.after(0, self._update_cache_label)
                self._log(
                    f"Auto-loaded: {recent_cache.name}  ({len(data)} unique names)", "ok")
        except Exception:
            pass  # Silently ignore if auto-load fails

    def _load_cache_file(self):
        initial = str(CACHE_DIR) if CACHE_DIR.exists() else str(APP_DIR)
        path = filedialog.askopenfilename(
            title="Select cache JSON",
            initialdir=initial,
            filetypes=[("JSON cache", "*.json"), ("All files", "*.*")])
        if not path:
            return
        try:
            data = load_cache(path)
            if data is None:
                raise ValueError("File could not be read.")
            self._cache_map  = data
            self._cache_path = path
            self._update_cache_label()
            self._log(
                f"Cache loaded: {os.path.basename(path)}  "
                f"({len(data)} unique names)", "ok")
        except Exception as exc:
            messagebox.showerror("Load error", str(exc))


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT

def main():
    if HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
        print("Tip: install tkinterdnd2 for drag-and-drop support.")

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure(
        "Vertical.TScrollbar",
        background=C["surface"], troughcolor=C["bg"],
        bordercolor=C["bg"], arrowcolor=C["muted"],
        relief="flat", gripcount=0,
    )

    PDFCrawler(root)
    root.mainloop()


if __name__ == "__main__":
    main()
